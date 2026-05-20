"""
Export Service - Generates Excel and PDF reports from project data.
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
from app.models.project_model import (
    get_all_projects, get_project, get_milestones,
    get_budget, get_risks, get_actions, get_notes
)


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ─────────────────────────────────────────────────────────────────────────────
# Excel Exports
# ─────────────────────────────────────────────────────────────────────────────

def export_all_projects_excel(output_path: str = None) -> str:
    """Export all projects to a formatted Excel file."""
    if not output_path:
        output_path = str(Path.home() / f"PMO_Portfolio_{_timestamp()}.xlsx")
    projects = get_all_projects()
    df = pd.DataFrame(projects)
    # Prettify columns
    rename = {
        "project_id": "Project ID", "name": "Project Name", "status": "Status",
        "phase": "Phase", "priority": "Priority", "manager": "Manager",
        "department": "Department", "client": "Client",
        "start_date": "Start Date", "end_date": "End Date",
        "progress": "Progress (%)", "description": "Description",
    }
    df = df.rename(columns=rename)
    df = df[[c for c in rename.values() if c in df.columns]]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Projects", index=False)
        ws = writer.sheets["Projects"]
        _auto_width(ws)
        _style_header(ws)
    return output_path


def export_project_detail_excel(project_id: str, output_path: str = None) -> str:
    """Export one project with all sub-tables to Excel."""
    if not output_path:
        output_path = str(Path.home() / f"Project_{project_id}_{_timestamp()}.xlsx")
    project = get_project(project_id)
    milestones = get_milestones(project_id)
    budget = get_budget(project_id)
    risks = get_risks(project_id)
    actions = get_actions(project_id)
    notes = get_notes(project_id)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary sheet
        summary_data = {k: [v] for k, v in (project or {}).items()}
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)
        _auto_width(writer.sheets["Summary"])
        _style_header(writer.sheets["Summary"])

        if milestones:
            pd.DataFrame(milestones).to_excel(writer, sheet_name="Milestones", index=False)
            _auto_width(writer.sheets["Milestones"])
            _style_header(writer.sheets["Milestones"])

        if budget:
            pd.DataFrame([budget]).to_excel(writer, sheet_name="Budget", index=False)
            _auto_width(writer.sheets["Budget"])
            _style_header(writer.sheets["Budget"])

        if risks:
            pd.DataFrame(risks).to_excel(writer, sheet_name="Risks", index=False)
            _auto_width(writer.sheets["Risks"])
            _style_header(writer.sheets["Risks"])

        if actions:
            pd.DataFrame(actions).to_excel(writer, sheet_name="Actions", index=False)
            _auto_width(writer.sheets["Actions"])
            _style_header(writer.sheets["Actions"])

        if notes:
            pd.DataFrame(notes).to_excel(writer, sheet_name="Notes", index=False)
            _auto_width(writer.sheets["Notes"])
            _style_header(writer.sheets["Notes"])

    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# PDF Export (optional)
# ─────────────────────────────────────────────────────────────────────────────

def export_project_pdf(project_id: str, output_path: str = None) -> str:
    """Export one project to a PDF summary report."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable)
    except ImportError:
        raise ImportError("reportlab is required for PDF export. Install via: pip install reportlab")

    if not output_path:
        output_path = str(Path.home() / f"Project_{project_id}_{_timestamp()}.pdf")

    project = get_project(project_id) or {}
    milestones = get_milestones(project_id)
    budget = get_budget(project_id) or {}
    risks = get_risks(project_id)
    actions = get_actions(project_id)

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    BLUE = colors.HexColor("#1565C0")
    LIGHT = colors.HexColor("#E3F2FD")

    title_style = ParagraphStyle("title", parent=styles["Title"],
                                  fontSize=18, textColor=BLUE, spaceAfter=6)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"],
                         fontSize=12, textColor=BLUE, spaceBefore=12)
    body = styles["BodyText"]

    story = []
    story.append(Paragraph(f"Project Report: {project.get('name', project_id)}", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", body))
    story.append(HRFlowable(width="100%", thickness=1, color=BLUE, spaceAfter=10))

    # General info table
    story.append(Paragraph("General Information", h2))
    info_data = [
        ["Project ID", project.get("project_id", ""), "Status", project.get("status", "")],
        ["Manager", project.get("manager", ""), "Phase", project.get("phase", "")],
        ["Client", project.get("client", ""), "Priority", project.get("priority", "")],
        ["Start Date", project.get("start_date", ""), "End Date", project.get("end_date", "")],
        ["Progress", f"{project.get('progress', 0)}%", "Department", project.get("department", "")],
    ]
    t = Table(info_data, colWidths=[4*cm, 5*cm, 4*cm, 4*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), LIGHT),
        ("BACKGROUND", (2, 0), (2, -1), LIGHT),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t)

    # Budget
    if budget:
        story.append(Paragraph("Budget & Financials", h2))
        bd = [
            ["Budget Type", budget.get("budget_type", ""), "Planned Budget", f"€{budget.get('planned_budget', 0):,.0f}"],
            ["Actual Cost", f"€{budget.get('actual_cost', 0):,.0f}", "Remaining", f"€{budget.get('remaining', 0):,.0f}"],
            ["Cost Variance", f"€{budget.get('cost_variance', 0):,.0f}", "", ""],
        ]
        bt = Table(bd, colWidths=[4*cm, 5*cm, 4*cm, 4*cm])
        bt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), LIGHT),
            ("BACKGROUND", (2, 0), (2, -1), LIGHT),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(bt)

    # Milestones
    if milestones:
        story.append(Paragraph("Milestones", h2))
        ms_data = [["Milestone", "Due Date", "Status", "Phase"]]
        for m in milestones:
            ms_data.append([m["name"], m.get("due_date", ""), m.get("status", ""), m.get("phase", "")])
        mt = Table(ms_data, colWidths=[7*cm, 3*cm, 3*cm, 4*cm])
        mt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(mt)

    # Risks
    if risks:
        story.append(Paragraph("Risks & Issues", h2))
        r_data = [["Description", "Impact", "Status", "Owner"]]
        for r in risks:
            r_data.append([r.get("description", ""), r.get("impact", ""), r.get("status", ""), r.get("owner", "")])
        rt = Table(r_data, colWidths=[7*cm, 3*cm, 3*cm, 4*cm])
        rt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(rt)

    doc.build(story)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _auto_width(ws):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _style_header(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="1565C0")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
